import telebot
from openpyxl import load_workbook

# 🔐 توكن البوت (من غير مسافات)
TOKEN = "8210513434:AAH_lK7WGqpHIMtbcUSFtDrzc07O0cA0pKU"

# 📁 مسار ملف الإكسيل
FILE_PATH = "data.xlsx"

# 🤖 تفعيل البوت
bot = telebot.TeleBot(TOKEN)

# 📊 تحميل ملف الإكسيل
try:
    wb = load_workbook(FILE_PATH, data_only=True)
    print("✅ تم تحميل ملف الإكسيل")
    print("📄 أسماء الشيتات:")
    for sheet in wb.sheetnames:
        print("-", sheet)
except Exception as e:
    print("❌ خطأ في ملف الإكسيل:", e)
    wb = None

# ▶️ أمر start
@bot.message_handler(commands=['start'])
def start(message):
    bot.reply_to(
        message,
        "👋 أهلاً بيك\n"
        "🔍 ابعت أي كلمة أو رقم للبحث في كل الشيتات"
    )

# 🔍 دالة البحث في كل الشيتات
def search_excel(term):
    results = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and term in str(cell):
                    results.append(f"📄 {sheet}\n🧾 {row}")
                    break

    return results

# 🤖 استقبال أي رسالة
@bot.message_handler(func=lambda m: True)
def handle_message(message):
    text = message.text.strip()

    if not text:
        bot.reply_to(message, "❌ ابعت كلمة للبحث")
        return

    results = search_excel(text)

    if results:
        reply = "\n\n".join(results[:5])  # أول 5 نتائج
    else:
        reply = "❌ مفيش نتائج"

    bot.reply_to(message, reply)

# ▶️ تشغيل البوت
print("🤖 البوت شغال...")
bot.infinity_polling()
